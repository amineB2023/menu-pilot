from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

from restaurant_bot.config import DATA_DIR
from restaurant_bot.database import get_connection
from restaurant_bot.services import menu_service


def period_range(kind: str) -> tuple[datetime, datetime, str]:
    now = datetime.now()
    if kind in {"today", "orders", "sales", "customers", "loyalty"}:
        start = datetime(now.year, now.month, now.day)
        end = start + timedelta(days=1)
        return start, end, start.strftime("%B %d, %Y")
    if kind == "week":
        start_date = date.today() - timedelta(days=date.today().weekday())
        start = datetime.combine(start_date, datetime.min.time())
        end = start + timedelta(days=7)
        return start, end, f"Week of {start.strftime('%B %d, %Y')}"
    start = datetime(now.year, now.month, 1)
    end = datetime(now.year + (now.month == 12), 1 if now.month == 12 else now.month + 1, 1)
    return start, end, start.strftime("%B %Y")


def rows_for_period(restaurant_id: int, start: datetime, end: datetime) -> list[dict]:
    with get_connection() as conn:
        order_rows = conn.execute(
            """
            SELECT o.*, u.username
            FROM orders o
            LEFT JOIN users u ON u.telegram_id = o.user_id
            WHERE o.restaurant_id = ?
              AND datetime(o.created_at) >= datetime(?)
              AND datetime(o.created_at) < datetime(?)
            ORDER BY o.created_at DESC, o.id DESC
            """,
            (restaurant_id, start.isoformat(" "), end.isoformat(" ")),
        ).fetchall()
        orders = []
        for row in order_rows:
            order = dict(row)
            items = conn.execute("SELECT * FROM order_items WHERE order_id = ? ORDER BY id", (order["id"],)).fetchall()
            order["items"] = [dict(item) for item in items]
            orders.append(order)
    return orders


def generate_xlsx_report(
    restaurant: dict,
    kind: str = "month",
    start: datetime | None = None,
    end: datetime | None = None,
    period_label: str | None = None,
) -> tuple[Path, str]:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Border, Font, PatternFill, Side

    if start is None or end is None or period_label is None:
        start, end, period_label = period_range(kind)
    orders = rows_for_period(int(restaurant["id"]), start, end)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"MenuPilot_Report_{period_label.replace(' ', '_').replace(',', '')}.xlsx"
    path = DATA_DIR / filename

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws) -> None:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        for column_cells in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 50)

    delivered = [o for o in orders if o["status"] == "delivered"]
    non_cancelled = [o for o in orders if o["status"] != "cancelled"]
    revenue = sum(int(o["subtotal_cents"]) for o in non_cancelled)
    khqr_revenue = sum(int(o["subtotal_cents"]) for o in non_cancelled if o["payment_method"] == "khqr" and o["payment_status"] == "paid")
    cash_revenue = sum(int(o["subtotal_cents"]) for o in non_cancelled if o["payment_method"] == "cash")
    customers = {o["user_id"] for o in orders}
    repeat_customers = sum(1 for user_id in customers if sum(1 for o in orders if o["user_id"] == user_id) > 1)
    rewards_redeemed = 0
    with get_connection() as conn:
        rewards_redeemed = int(
            conn.execute(
                """
                SELECT COUNT(*) FROM reward_redemptions
                WHERE restaurant_id = ?
                  AND datetime(redeemed_at) >= datetime(?)
                  AND datetime(redeemed_at) < datetime(?)
                """,
                (restaurant["id"], start.isoformat(" "), end.isoformat(" ")),
            ).fetchone()[0]
        )

    overview.append(["Restaurant", restaurant["name"]])
    overview.append(["Period", period_label])
    overview.append(["Generated date", datetime.now().strftime("%Y-%m-%d %H:%M")])
    overview.append([])
    overview.append(["Metric", "Value"])
    metrics = [
        ("Total Orders", len(orders)),
        ("Delivered Orders", len(delivered)),
        ("Cancelled Orders", sum(1 for o in orders if o["status"] == "cancelled")),
        ("Revenue", menu_service.cents_to_usd(revenue, restaurant["currency_symbol"])),
        ("KHQR Revenue", menu_service.cents_to_usd(khqr_revenue, restaurant["currency_symbol"])),
        ("Cash Revenue", menu_service.cents_to_usd(cash_revenue, restaurant["currency_symbol"])),
        ("Average Order Value", menu_service.cents_to_usd(revenue // max(len(non_cancelled), 1), restaurant["currency_symbol"])),
        ("Total Customers", len(customers)),
        ("Repeat Customers", repeat_customers),
        ("Rewards Redeemed", rewards_redeemed),
    ]
    for metric in metrics:
        overview.append(metric)
    for cell in overview["A"]:
        cell.font = bold

    orders_ws = wb.create_sheet("Orders")
    orders_ws.append(["Order ID", "Date", "Customer Name", "Phone", "Delivery/Pickup", "Payment Method", "Payment Status", "Order Status", "Items", "Quantity", "Subtotal", "Discount", "Final Total", "Notes"])
    for order in orders:
        item_names = ", ".join(f"{item['quantity']}x {item['item_name']}" for item in order["items"])
        qty = sum(int(item["quantity"]) for item in order["items"])
        subtotal = menu_service.cents_to_usd(int(order["subtotal_cents"]), restaurant["currency_symbol"])
        orders_ws.append([order["order_code"], order["created_at"], order.get("customer_name"), order["phone"], order["fulfillment_type"], order["payment_method"], order["payment_status"], order["status"], item_names, qty, subtotal, 0, subtotal, order.get("notes") or ""])

    best_ws = wb.create_sheet("Best Sellers")
    best_ws.append(["Item", "Total Orders", "Quantity Sold", "Revenue Generated"])
    best: dict[str, dict[str, int]] = {}
    for order in non_cancelled:
        for item in order["items"]:
            bucket = best.setdefault(item["item_name"], {"orders": 0, "qty": 0, "revenue": 0})
            bucket["orders"] += 1
            bucket["qty"] += int(item["quantity"])
            bucket["revenue"] += int(item["line_total_cents"])
    for name, data in sorted(best.items(), key=lambda pair: pair[1]["qty"], reverse=True):
        best_ws.append([name, data["orders"], data["qty"], menu_service.cents_to_usd(data["revenue"], restaurant["currency_symbol"])])

    customers_ws = wb.create_sheet("Customers")
    customers_ws.append(["Customer Name", "Telegram Username", "Phone", "Orders Count", "Total Spent", "Average Spend", "Last Order", "Favorite Item", "Loyalty Points", "Rewards Redeemed"])
    for user_id in customers:
        user_orders = [o for o in orders if o["user_id"] == user_id]
        spent = sum(int(o["subtotal_cents"]) for o in user_orders if o["status"] != "cancelled")
        item_counts: dict[str, int] = {}
        for order in user_orders:
            for item in order["items"]:
                item_counts[item["item_name"]] = item_counts.get(item["item_name"], 0) + int(item["quantity"])
        favorite = max(item_counts, key=item_counts.get) if item_counts else ""
        latest = max(user_orders, key=lambda o: o["created_at"])
        with get_connection() as conn:
            points_row = conn.execute("SELECT points FROM user_loyalty_points WHERE restaurant_id = ? AND user_id = ?", (restaurant["id"], user_id)).fetchone()
            reward_count = conn.execute("SELECT COUNT(*) FROM reward_redemptions WHERE restaurant_id = ? AND user_id = ?", (restaurant["id"], user_id)).fetchone()[0]
        customers_ws.append([latest.get("customer_name"), latest.get("username"), latest.get("phone"), len(user_orders), menu_service.cents_to_usd(spent, restaurant["currency_symbol"]), menu_service.cents_to_usd(spent // max(len(user_orders), 1), restaurant["currency_symbol"]), latest["created_at"], favorite, int(points_row["points"]) if points_row else 0, int(reward_count)])
    if customers_ws.max_row > 1:
        customers_ws["A2"].fill = PatternFill("solid", fgColor="FFF2CC")

    analytics_ws = wb.create_sheet("Sales Analytics")
    analytics_ws.append(["Metric", "Value"])
    analytics = [
        ("Daily Revenue", menu_service.cents_to_usd(revenue, restaurant["currency_symbol"])),
        ("Weekly Revenue", menu_service.cents_to_usd(revenue, restaurant["currency_symbol"])),
        ("Monthly Revenue", menu_service.cents_to_usd(revenue, restaurant["currency_symbol"])),
        ("KHQR Orders", sum(1 for o in orders if o["payment_method"] == "khqr")),
        ("Cash Orders", sum(1 for o in orders if o["payment_method"] == "cash")),
        ("Average Basket Size", round(sum(sum(int(i["quantity"]) for i in o["items"]) for o in orders) / max(len(orders), 1), 2)),
        ("Peak Order Hour", max((datetime.fromisoformat(o["created_at"]).hour for o in orders), default="")),
    ]
    for row in analytics:
        analytics_ws.append(row)
    chart = BarChart()
    chart.title = "Payment Split"
    data = Reference(analytics_ws, min_col=2, min_row=5, max_row=6)
    cats = Reference(analytics_ws, min_col=1, min_row=5, max_row=6)
    chart.add_data(data)
    chart.set_categories(cats)
    analytics_ws.add_chart(chart, "D2")

    loyalty_ws = wb.create_sheet("Loyalty")
    loyalty_ws.append(["Customer", "Points Earned", "Points Spent", "Current Balance", "Rewards Redeemed"])
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT u.full_name,
                   COALESCE(SUM(o.loyalty_points_awarded), 0) AS earned,
                   COALESCE(SUM(rr.points_spent), 0) AS spent,
                   COALESCE(ulp.points, 0) AS balance,
                   COUNT(rr.id) AS redeemed
            FROM users u
            LEFT JOIN orders o ON o.user_id = u.telegram_id AND o.restaurant_id = ?
            LEFT JOIN reward_redemptions rr ON rr.user_id = u.telegram_id AND rr.restaurant_id = ?
            LEFT JOIN user_loyalty_points ulp ON ulp.user_id = u.telegram_id AND ulp.restaurant_id = ?
            GROUP BY u.telegram_id
            HAVING earned > 0 OR spent > 0 OR balance > 0 OR redeemed > 0
            """,
            (restaurant["id"], restaurant["id"], restaurant["id"]),
        ).fetchall()
    for row in rows:
        loyalty_ws.append([row["full_name"], row["earned"], row["spent"], row["balance"], row["redeemed"]])

    for ws in wb.worksheets:
        style_sheet(ws)
    wb.save(path)
    return path, period_label
